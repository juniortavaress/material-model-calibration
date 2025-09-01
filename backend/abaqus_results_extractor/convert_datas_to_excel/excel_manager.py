import os
import yaml
import math
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from typing import Dict, List, Tuple, Union, Optional
import re
from backend.config.yaml_manager import YamlManager

class ExcelManager():
    def create_excel_results(self, forces: bool, temperature: bool, chip: bool) -> None:
        """
        Creates an Excel file with two sheets ('Data' and 'Info') containing columns
        defined based on the YAML configuration. Skips creation if the file already exists.

        Args:
            forces (bool): Include force-related columns?
            temperature (bool): Include temperature-related columns?
            chip (bool): Include chip-related columns?
        """
        excel_results_path = os.path.join(self.excel_files, "datas.xlsx")
        if os.path.exists(excel_results_path):
            return 
        
        if self.execution_mode != "standalone":
            # Load parameters from YAML configuration file
            yaml_info_data = YamlManager.load_yaml(self, self.yaml_project_info)

            # Define columns for the main data sheet
            param_selection = yaml_info_data.get("07. Parameters to Iterate", {})
            parameters_list = [key for key, value in param_selection.items() if value]
            fixed_cols = ["Iteration Number", "Parameter Set", "Best Set of Iteration", "Condition"]
            error = ["Error"]
            forces_cols = ["Error Fc", "Error Fn"] if forces else []
            temp_cols = ["Error Temp"] if temperature else []
            chip_cols = ["Error CCR", "Error CSR"] if chip else []
            param_cols = [f"Parameter {p}" for p in parameters_list] 
            columns = fixed_cols + param_cols + error + forces_cols + temp_cols + chip_cols
            df_data = pd.DataFrame(0, columns=columns, index=range(1))

            # Prepare condition information for the "Info" sheet
            data_rows = []
            for value in yaml_info_data.get("05. Conditions", {}).values():
                if isinstance(value, dict):
                    cond = value["Cutting Properties"]["name"]
                    v = value["Cutting Properties"]["velocity"]
                    h = value["Cutting Properties"]["deepCuth"]
                    gam = value["Cutting Properties"]["rakeAngle"]
                    data_rows.append({"Condition": cond,"Velocity": v,"Deep of Cuth": h,"Rake Angle": gam})
            df_info = pd.DataFrame(data_rows)

            # Save both DataFrames into a single Excel file with two separate sheets
            excel_results_path = os.path.join(self.excel_files, "datas.xlsx")
            with pd.ExcelWriter(excel_results_path, engine="openpyxl") as writer:
                df_data.to_excel(writer, sheet_name="Data", index=False)
                df_info.to_excel(writer, sheet_name="Info", index=False)
        else:
            fixed_cols = ["Filename", "Error"]
            forces_cols = ["Error Fc", "Error Fn"] if forces else []
            temp_cols = ["Error Temp"] if temperature else []
            chip_cols = ["Error CCR", "Error CSR"] if chip else []
            columns = fixed_cols + forces_cols + temp_cols + chip_cols
            df_data = pd.DataFrame(0, columns=columns, index=range(1))

            excel_results_path = os.path.join(self.excel_files, "datas.xlsx")
            with pd.ExcelWriter(excel_results_path, engine="openpyxl") as writer:
                df_data.to_excel(writer, sheet_name="Data", index=False)


    def create_datas(self, filename: str, forces_summary: Dict, temperature_summary: Dict, chip_summary: Dict) -> None:
        """
        Calculates simulation errors vs experimental values and writes results to Excel.

        Args:
            filename (str): Simulation filename.
            forces_summary (dict): Simulated force results.
            temperature_summary (dict): Simulated temperature results.
            chip_summary (dict): Simulated chip geometry results.
        """
        error_fc, error_fn, error_temp, error_ccr, error_csr = None, None, None, None, None
        iteration_number, set_number, condition = ExcelManager._get_file_info(self, filename)
        target_values = ExcelManager._get_target_values(self, condition)

        if forces_summary: 
            simulated_cutting_forces = forces_summary[filename]["Cutting Force [N].mean"]
            simulated_normal_forces = forces_summary[filename]["Normal Force [N].mean"]
            error_fc = (simulated_cutting_forces - target_values[0])/target_values[0]
            error_fn = (simulated_normal_forces - target_values[1])/target_values[1]
        
        if temperature_summary:
            simulated_temperature = temperature_summary[filename]["Maximum Temperature at Last Frame [°C]"]
            error_temp = (simulated_temperature - target_values[2])/target_values[2]

        if chip_summary:
            chip_compression_ratio = chip_summary[filename]["Chip Compression Ratio (CCR)"]
            chip_segmentation_ratio = chip_summary[filename]["Chip Segmentatio Ratio (CSR)"]
            error_ccr = (chip_compression_ratio - target_values[3])/target_values[3] if chip_compression_ratio is not None else None
            error_csr = (chip_segmentation_ratio - target_values[4])/target_values[4] if chip_segmentation_ratio is not None else None
    
        if self.execution_mode != "standalone":
            param_set = ExcelManager._get_param_values(self, iteration_number, set_number)
            data = {"Iteration Number": int(iteration_number), "Parameter Set": int(set_number),"Condition": condition[4:]}
            for key, value in param_set.items():
                data[f"Parameter {key}"] = value          
        else:
            data = {"Filename": str(filename)}

        error_values = {}
        if forces_summary:
            error_values["Error Fc"] = error_fc 
            error_values["Error Fn"] = error_fn 
        if temperature_summary:
            error_values["Error Temp"] = error_temp 
        if chip_summary:
            error_values["Error CCR"] = error_ccr 
            error_values["Error CSR"] = error_csr
        
        if chip_summary and (error_ccr is None or error_csr is None):
            combined_error = None
        elif temperature_summary and error_temp is None:
            combined_error = None
        elif forces_summary and (error_fn is None or error_fc is None):
            combined_error = None
        else:
            efc = error_fc if error_fc is not None else 0
            efn = error_fn if error_fn is not None else 0
            et = error_temp if error_temp is not None else 0
            eccr = error_ccr if error_ccr is not None else 0
            ecsr = error_csr if error_csr is not None else 0
            combined_error = math.sqrt((self.wfc * efc ** 2) + (self.wfn * efn ** 2) + (self.wt * et ** 2) + (self.wccr * eccr ** 2) + (self.wcsr * ecsr ** 2))

        error_values["Error"] = combined_error
        data.update(error_values)
        ExcelManager._save_iteration(self, data)


    def _save_iteration(self, data: Dict[str, Union[str, int, float]]) -> None:
        """
        Appends a new row to the Excel file and identifies the best parameter set per iteration.

        Args:
            data (dict): New row of results.
        """
        new_info = pd.DataFrame([data])
        data_path = os.path.join(self.excel_files, "datas.xlsx")
        old_df = pd.read_excel(data_path, engine="openpyxl")
        new_df = pd.concat([old_df, new_info], ignore_index=True) 
        new_df = new_df.loc[~(new_df == 0).all(axis=1)]  

        if self.execution_mode != "standalone":
            last_iteration = new_df["Iteration Number"].max()
            df_last_iteration = new_df[new_df["Iteration Number"] == last_iteration]

            valid_groups = df_last_iteration.groupby("Parameter Set")["Error"].apply(lambda x: x.notna().all())
            valid_parameter_sets = valid_groups[valid_groups].index
            df_valid = df_last_iteration[df_last_iteration["Parameter Set"].isin(valid_parameter_sets)]
            df_avg_errors_last_iter = df_valid.groupby("Parameter Set")["Error"].mean().reset_index()

            best_row_last_iter = df_avg_errors_last_iter.loc[df_avg_errors_last_iter["Error"].idxmin()]
            best_parameter_set_last_iter = best_row_last_iter["Parameter Set"]

            new_df.loc[new_df["Iteration Number"] == last_iteration, "Best Set of Iteration"] = best_parameter_set_last_iter
            columns = list(new_df.columns)
            columns.insert(2, columns.pop(columns.index("Best Set of Iteration"))) 
            new_df = new_df[columns] 

        with pd.ExcelWriter(data_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            new_df.to_excel(writer, sheet_name="Data", index=False)
        ExcelManager._format_file(self, data_path)


    def _format_file(self, data_path: str) -> None:
        """
        Applies formatting to the Excel file: column width, centering, bold headers, and table style.

        Args:
            data_path (str): Path to the Excel file.
        """
        wb = openpyxl.load_workbook(data_path)  
        for i, ws in enumerate(wb.worksheets):
            bold_font = Font(bold=True)
            for cell in ws[1]: 
                cell.font = bold_font
            ws.sheet_view.showGridLines = False

            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter 

                for cell in col:
                    if cell.value:
                        cell.alignment = Alignment(horizontal="center", vertical="center")  
                        max_length = max(max_length, len(str(cell.value)))  
                ws.column_dimensions[col_letter].width = max_length + 2 

            for col in ws.columns:
                for cell in col:
                    if cell.value and isinstance(cell.value, (int, float)):
                        if ws.cell(row=1, column=cell.column).value.startswith("Error"):
                            cell.number_format = '0.00%'

            existing_table_names = set(ws.tables.keys())
            table_name = f"ResultsTable_{i}"
            if table_name not in existing_table_names:
                table_range = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"
                table = Table(displayName=table_name, ref=table_range)
                style = TableStyleInfo(showFirstColumn=False,showLastColumn=False,showRowStripes =True,showColumnStripes=True)
                table.tableStyleInfo = style
                ws.add_table(table)
            wb.save(data_path)
            wb.close()


    def _get_file_info(self, filename: str) -> tuple[str, str, str]:
        """
        Extracts iteration number, parameter set, and condition from a filename.

        Args:
            filename (str): Input filename.

        Returns:
            Tuple[str, str, str]: (iteration_number, set_number, condition)
        """
        if self.execution_mode != "standalone":
            condition = filename.split("_it_")[0]
            condition = filename.split("_it_")[0]
            iteration_number = filename.split("_it_")[1].split("_set_")[0]
            set_number = filename.split("_set_")[1]
            return iteration_number, set_number, condition
        else:
            match = re.search(r'cond\d{2}', filename)
            return None, None, match.group(0) if match else None
        

    def _get_param_values(self, iteration_number: str, set_number: str) -> Dict[str, Union[int, float]]:
        """
        Retrieves parameter values for a given iteration and parameter set.

        Args:
            iteration_number (str): Iteration number.
            set_number (str): Parameter set number.

        Returns:
            Dict[str, Union[int, float]]: Dictionary of parameter values.
        """
        data = YamlManager.load_yaml(self, self.yaml_parameters)
        for iteration, values in data.items():
            if iteration[-2:] == iteration_number:
                for set, param_values in values.items():
                    if set[-2:] == set_number:
                        param_set = param_values["Parameters"]
        return param_set


    def _get_target_values(self, condition: str) -> List[Optional[float]]:
        """
        Retrieves experimental target values for a given condition.

        Args:
            condition (str): Condition name.

        Returns:
            List[Optional[float]]: [fc, fn, temp, ccr, csr]
        """
        target_values={}
        data = YamlManager.load_yaml(self, self.yaml_project_info)

        # Process cutting conditions and experimental data
        if "05. Conditions" in data:
            conditions_info = data["05. Conditions"]
            for _, value in conditions_info.items():
                if isinstance(value, dict):
                    for info, info_value in value.items():
                        if info == "Cutting Properties":
                            name = info_value['name']
                        elif info == "Experimental Datas":
                            fields = {"cutting_force": "fc", "normal_force": "fn", "temp": "temp", "chip_compression": "ccr", "chip_segmentation": "csr"}
                            values = {alias: float(info_value[key]) for key, alias in fields.items() if key in info_value}
                            target_values[name] = values
            
            if self.execution_mode != "standalone":
                condition = condition[4:]

            if condition in target_values:
                target_cutting_force = target_values[condition].get("fc", None)
                target_normal_force = target_values[condition].get("fn", None)
                target_temp = target_values[condition].get("temp", None)
                target_chip_compression_ratio = target_values[condition].get("ccr", None)
                target_chip_segentation_ratio = target_values[condition].get("csr", None)
        return [target_cutting_force, target_normal_force, target_temp, target_chip_compression_ratio, target_chip_segentation_ratio]
        

