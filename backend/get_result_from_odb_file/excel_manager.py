import os
import yaml
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

class ExcelManager():
    def create_excel_results(self):
        # Caminho do arquivo Excel
        data_path = os.path.join(self.excel_files, "datas.xlsx")

        # Obter nomes dos parâmetros
        yaml_path = os.path.join(self.user_config, "parameters.yaml")
        with open(yaml_path, "r", encoding="utf-8") as file:
            data = yaml.safe_load(file)

        param_set = next((param_values["Parameters"] for values in data.values() for param_values in values.values()))
        parameter_names = list(param_set.keys())

        # Criar DataFrame de Dados se o arquivo não existir
        if not os.path.exists(data_path):
            columns = ["Iteration Number", "Parameter Set", "Best Set of Iteration", "Condition", "Error", "Error Fc", "Error Fn", "Error CCR", "Error CSR"]
            parameter_columns = [f"Parameter {key}" for key in parameter_names]
            columns = columns[:4] + parameter_columns + columns[4:]

            df_data = pd.DataFrame(0, columns=columns, index=range(1))
            # df_data.index.name = "Simulation"
        else:
            df_data = pd.read_excel(data_path, sheet_name="Data", engine="openpyxl")



        # Carregar informações do YAML
        formatted_names = {}
        with open(self.project_infos_path, "r", encoding="utf-8") as file:
            cond_data = yaml.safe_load(file)

        for value in cond_data.get("3. Conditions", {}).values():
            cond = value["Cutting Properties"]["name"]
            v = value["Cutting Properties"]["velocity"]
            h = value["Cutting Properties"]["deepCuth"]
            gam = value["Cutting Properties"]["rakeAngle"]

            formatted_name = f"v{v}_h{h}_gam{gam}"
            formatted_names[cond] = formatted_name

        df_info = pd.DataFrame(list(formatted_names.items()), columns=["Condition", "Values"])

        # Salvar no arquivo Excel com múltiplas planilhas
        with pd.ExcelWriter(data_path, engine="openpyxl") as writer:
            df_data.to_excel(writer, sheet_name="Data", index=False)
            df_info.to_excel(writer, sheet_name="Info", index=False)




    def create_datas(self):
        """
        Extracts results from the simulation output files and converts them into a structured format.
        """
        for filename, force_and_temp_datas in self.force_and_temp_datas.items():
            file = filename
            chip_datas = self.chip_datas[filename]

            info = file.split("_int")
            condition = file.split("_int_")[0]
            param_info = f"int{info[1]}"
            iteration_number = (param_info.split("int_")[1]).split("_set_")[0]
            set_number = (param_info.split("int_")[1]).split("_set_")[1]

            simulated_cutting_forces = force_and_temp_datas["Cutting Force [N].mean"]
            simulated_normal_forces = force_and_temp_datas["Normal Force [N].mean"]
            simulated_temperature = force_and_temp_datas["Maximum Temperature at Last Frame [°C]"]
            chip_compression_ratio = chip_datas["Chip Compression Ratio (CCR)"]
            chip_segmentation_ratio = chip_datas["Chip Segmentatio Ratio (CSR)"]

            yaml_path = os.path.join(self.user_config, "parameters.yaml")
            with open(yaml_path, "r", encoding="utf-8") as file:
                data = yaml.safe_load(file)

            for iteration, values in data.items():
                if iteration[-2:] == iteration_number:
                    for set, param_values in values.items():
                        if set[-2:] == set_number:
                            param_set = param_values["Parameters"]

            if condition[4:] in self.target_values:
                target_cutting_force = self.target_values[condition[4:]].get("fc")
                target_normal_force = self.target_values[condition[4:]].get("fn")
                target_chip_compression_ratio = self.target_values[condition[4:]].get("ccr")
                target_chip_segentatio_ratio = self.target_values[condition[4:]].get("csr")

            error_fc = (simulated_cutting_forces - target_cutting_force)/target_cutting_force
            error_fn = (simulated_normal_forces - target_normal_force)/target_normal_force
            error_ccr = (chip_compression_ratio - target_chip_compression_ratio)/target_chip_compression_ratio
            error_csr = (chip_segmentation_ratio - target_chip_segentatio_ratio)/target_chip_segentatio_ratio
            error = (0.5 * abs(error_fc)) + (0.1 * abs(error_fn)) + (0.2 * abs(error_ccr)) + (0.2 * abs(error_csr))

            data = {"Iteration Number": int(iteration_number), "Parameter Set": int(set_number),"Condition": condition[4:]}

            for key, value in param_set.items():
                data[f"Parameter {key}"] = value

            data.update({
                "Error": error,
                "Error Fc": error_fc,
                "Error Fn": error_fn,
                "Error CCR": error_ccr,
                "Error CSR": error_csr})
            
            new_info = pd.DataFrame([data])
            data_path = os.path.join(self.excel_files, "datas.xlsx")
            old_df = pd.read_excel(data_path, engine="openpyxl")
            new_df = pd.concat([old_df, new_info], ignore_index=True)  # Sem ignore_index para manter o índice original
            new_df = new_df.loc[~(new_df == 0).all(axis=1)]  

            last_iteration = new_df["Iteration Number"].max()
            df_last_iteration = new_df[new_df["Iteration Number"] == last_iteration]
            best_row = df_last_iteration.loc[df_last_iteration["Error"].idxmin()]
            best_parameter_set = best_row["Parameter Set"]
            
            new_df.loc[new_df["Iteration Number"] == last_iteration, "Best Set of Iteration"] = best_parameter_set
            columns = list(new_df.columns)
            columns.insert(2, columns.pop(columns.index("Best Set of Iteration"))) 
            new_df = new_df[columns] 

            with pd.ExcelWriter(data_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                new_df.to_excel(writer, sheet_name="Data", index=False)

            ExcelManager.format_file(self, data_path)

    def format_file(self, data_path):
        # Ajustar formatação do Excel
        wb = openpyxl.load_workbook(data_path)  # Abrir o arquivo Excel
        

        for i, ws in enumerate(wb.worksheets):
            # Ajustar largura das colunas e centralizar o texto
            bold_font = Font(bold=True)
            for cell in ws[1]:  # Primeira linha (cabeçalho)
                cell.font = bold_font

            ws.sheet_view.showGridLines = False

            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter  # Letra da coluna no Excel

                for cell in col:
                    if cell.value:
                        cell.alignment = Alignment(horizontal="center", vertical="center")  # Centralizar texto
                        max_length = max(max_length, len(str(cell.value)))  # Verificar maior tamanho do texto

                ws.column_dimensions[col_letter].width = max_length + 2  # Ajustar largura da coluna

            for col in ws.columns:
                for cell in col:
                    if cell.value and isinstance(cell.value, (int, float)):
                        # Verificar se o nome da coluna começa com "Error"
                        if ws.cell(row=1, column=cell.column).value.startswith("Error"):
                            # Aplicar formato de porcentagem para as células de erro
                            cell.number_format = '0.00%'

            existing_table_names = set(ws.tables.keys())
            if self.number_of_iterations == 1:
                table_name = f"ResultsTable_{i}"
                
                if table_name not in existing_table_names:
                    table_range = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"
                    table = Table(displayName=table_name, ref=table_range)

                    style = TableStyleInfo(
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=True
                    )
                    table.tableStyleInfo = style
                    ws.add_table(table)

            # Salvar as modificações
            wb.save(data_path)
            wb.close()
